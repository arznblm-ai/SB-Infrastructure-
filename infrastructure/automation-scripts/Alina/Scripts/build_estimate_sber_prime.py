"""Смета «Сбер Прайм x ВиТ» (тендер Mosaic, 18.08.2026) в xlsx-шаблоне Антона.

Шаблон: копия EST_Зеленая Линия_Ai_15sec_14-08-26.xlsx (Drive id 1Uubr-vvVnMD7xbj1PF0rN5g6miNE4aP8),
структура MAIN <- PRE/PROD/POST <- EXPENSES. РФ-клиент: двуязычные подписи шаблона оставлены, валюта - RUB
(колонка USD по курсу на MAIN!F27 остаётся справочно).
Ставки EXPENSES = карта Антона 18.08 (маржа внутри). Fee / налоги - инпуты D22/D23, по умолчанию 0%
(правило из памяти 18.08; в шаблоне Зелёной Линии стояли 10% / 15%). Строки без ставки (раскадровка,
диктор, композитор) оставлены с пустой ставкой и пометкой «?» - итог по ним 0, смета неполная.

Использование:
    python3 build_estimate_sber_prime.py <template.xlsx> <out.xlsx>
    python3 inject_values.py <out.xlsx> <out_cached.xlsx>   # кэш значений формул для Quick Look
Вариант в файле - A-lite (см. брейкдаун, секции 5-6); полный A - добавить дни: CG sup 2, AI artist +1,
3D designer +1, Animator +1, Producer +2, AD +1, Rigger +1, Modeller +1.
"""
import sys
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell

SRC = sys.argv[1] if len(sys.argv) > 1 else "template.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "EST_SberPrime_ViT_CG-AI_15sec_18-08-26.xlsx"
FX = 85.0136  # ЦБ РФ USD/RUB на 18.08.2026 (cbr.ru, XML_daily) - справочно

wb = load_workbook(SRC)

# ---------------- EXPENSES: карта Антона 18.08 ----------------
ex = wb["EXPENSES"]
card = {8: 15000, 9: 25000, 10: 30000, 11: 20000, 12: 25000,                        # main crew
        14: 25000, 15: 15000, 16: 20000, 17: 20000, 18: 25000, 19: 15000, 20: 25000, 21: 20000,  # 3D
        23: 15000, 24: 20000, 25: 20000, 26: 20000, 27: 15000, 28: 18000, 29: 15000, 30: 30000}  # 2D & post
for r, v in card.items():
    ex.cell(row=r, column=5).value = v
    ex.cell(row=r, column=5).number_format = "#,##0"
ex["B32"] = "Ставки: рейт-карта студии от 18.08.2026 (RUB в день, маржа внутри). Sound producer / диктор / аудиостоки / ресайзы - ставки листа POST PRODUCTION (шаблон 14.08)."
ex["B32"].font = Font(name="Roboto", italic=True, sz=9)


def unmerge_cd(ws):
    for rng in list(ws.merged_cells.ranges):
        if rng.min_col == 3 and rng.max_col == 4 and rng.min_row == rng.max_row:
            ws.unmerge_cells(str(rng))


def reset_inputs(ws):
    """Обнулить количество/дни шаблона (там остались значения Зелёной Линии)."""
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        for c in row:
            if isinstance(c, MergedCell):
                continue
            if c.column in (5, 8) and c.row >= 6 and not (isinstance(c.value, str) and c.value.startswith("=")):
                if c.column == 5 and isinstance(c.value, str) and c.value.strip() == "":
                    c.value = None
                elif isinstance(c.value, (int, float)):
                    c.value = None
            if c.column == 11:
                c.number_format = "#,##0"
            if c.column == 7:
                c.number_format = "#,##0"


def line(ws, r, note=None, qty=None, days=None, label=None, rate=None):
    if label is not None:
        ws.cell(row=r, column=3).value = label
    if note is not None:
        ws.cell(row=r, column=4).value = note
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=4).font = Font(name="Roboto", sz=9)
    if qty is not None:
        ws.cell(row=r, column=5).value = qty
    if days is not None:
        ws.cell(row=r, column=8).value = days
    if rate is not None:
        ws.cell(row=r, column=7).value = rate


# ---------------- PRE PRODUCTION ----------------
pp = wb["PRE PRODUCTION"]; unmerge_cd(pp); reset_inputs(pp)
pp["B16"] = "Preproduction expenses / Предпроизводственные затраты"; pp["B22"] = "Tests / Тесты"  # в шаблоне =MAIN!AI10 (пусто)
line(pp, 8, "координация 13 дней с выходными (20.08-01.09), связь 24/7, показы 21.08 / 26.08 / 29.08 / 31.08, приёмка, сдача 01.09", 1, 7)
line(pp, 9, "стиль-фреймы ×4 в пиксар-стиле, лук, ревью анимации и v1", 1, 5)
line(pp, 10, "не в A-lite; опция: пайп под 13 дней (риги клиента -> Blender, рендер-сетап) +2 дня", None, None)
line(pp, 11, "стиль-фреймы (штаб с героями, экран, люк + лента, коробка + пэкшот)", 1, 1)
line(pp, 17, "н/п - сценарий и идея от клиента (подтверждён 18.08)")
line(pp, 18, "? - раскадровщица агентства (брифуют 18.08), кост уточнить у агентства", 1, 1)
line(pp, 23, "н/п - модели КосмоПрайм, 3D-лента Прайм, развёртка коробки - от клиента")
line(pp, 24, "тест ригов и рендера - внутри 3D designer")
pp["E25"] = None

# ---------------- PRODUCTION ----------------
pr = wb["PRODUCTION"]; unmerge_cd(pr); reset_inputs(pr)
pr["B7"] = "3D scene assembly / Сборка 3D-сцен и анимация (Blender)"
line(pr, 8, "окружение штаба (стилизованное), люк, стол, лента Прайм из 3D клиента, шейдинг «пиксар», свет, рендер, пэкшот-сцена", 1, 8)
line(pr, 9, "коробка КидсКомбо из развёртки, пропсы штаба", 1, 1)
line(pr, 13, "проверка / адаптация ригов героев клиента; без ригов от клиента - 4-5 дней", 1, 1)
line(pr, 14, "5 сцен × 3 героя: занятия, реакция на сигнал, бег, прыжок и слайд по ленте, скат в коробку", 1, 6)
line(pr, 15, "внутри 3D designer")
line(pr, 20, "н/п - полный CG, без съёмки")

# ---------------- POST PRODUCTION ----------------
po = wb["POST PRODUCTION"]; unmerge_cd(po); reset_inputs(po)
po["B7"] = "TC & Edit / Цветокоррекция и монтаж"; po["B15"] = "CGI & VFX / Графика и композ"; po["B25"] = "Sound / Звук"  # были =MAIN!AI15/17/18
line(po, 8, "аниматик 15 сек с VO-болванкой, сборка v1 / v2, мастер ТВ", 1, 2)
line(po, 9, "внутри композа")
line(po, 11, "адаптация OOH 5 сек (формат уточнить; вертикаль/квадрат = перекадровка +1 день 3D)", 1, 1, rate=70000)
line(po, 18, "композ рендер-пассов, экран КосмоПрайм, юр. плашки, титры, лого; цветокор внутри", 1, 3)
line(po, 26, "сведение, саунд-дизайн, мастер звука для ТВ", 1, 2)
line(po, 27, "? - диктор + запись, права ТВ РФ + интернет мир 01.09-05.10; ставка шаблона 50 000/проект - ориентир, подтвердить", 1, 1, rate=None)
po["G27"] = None
line(po, 28, "н/п - по брифу композитор, не сток")
line(po, 29, "? - оригинальная музыка + права ТВ РФ + интернет мир 01.09-05.10; ставки нет", 1, 1,
     label="Composer (music) / Композитор - музыка и права")
po["F29"] = "per proj"
po["K29"] = "=E29*(G29*H29+I29*J29)"
po["K29"].number_format = "#,##0"
for src, dst in (("K28", "K29"), ("F28", "F29"), ("E28", "E29"), ("H28", "H29")):
    po[dst].border = copy(po[src].border); po[dst].font = copy(po[src].font); po[dst].alignment = copy(po[src].alignment)

# ---------------- MAIN ----------------
m = wb["MAIN"]
m["C6"] = "СберПрайм х Вкусно и точка"
m["C7"] = "CG/AI ролик «КосмоПрайм» 15 сек + адаптация OOH 5 сек"
m["C8"] = "Mosaic (тендер, смета от 19.08.2026)"
# v1.1: комиссия 5% и налоги 15% - как в итоговой смете Антона VTB Privilege 18.08 (РФ-клиент, карта 18.08); инпуты
m["D22"] = 0.05; m["E22"] = "=E21*D22"; m["D22"].number_format = "0.0%"
m["D23"] = 0.15; m["E23"] = "=(E21+E22)*D23"; m["D23"].number_format = "0%"
m["E24"] = "=E21+E22+E23"
m["F27"] = FX; m["F27"].number_format = "#,##0.00"
m["E27"] = "Курс USD/RUB (ЦБ РФ, 18.08.2026), справочно"
for r in range(11, 25):
    if not isinstance(m.cell(row=r, column=5), MergedCell):
        m.cell(row=r, column=5).number_format = "#,##0"
    if not isinstance(m.cell(row=r, column=6), MergedCell):
        m.cell(row=r, column=6).number_format = "#,##0.00"

notes = [
    ("Примечания", True),
    ("Скоуп: 1 CG/AI ролик 15 сек (ТВ РФ + интернет мир, права 01.09-05.10.2026) + 1 адаптация OOH 5 сек; сценарий «Идея 2» клиента; стиль детский, а-ля пиксар. Вариант A-lite: герои - 3D-модели клиента, анимация в Blender.", False),
    ("Клиент передаёт: модели КосмоПрайм (формат и риги уточняются), 3D-лента Прайм, развёртка коробки КидсКомбо, логотипы и юр. тексты для пэкшота и плашек.", False),
    ("Процесс: показ №1 стиль-фреймы + аниматик + голос (21.08) → №2 черновая анимация + демо музыки (26.08) → v1 (29.08) → правки, 1 раунд включён → v2 (31.08) → мастер (01.09). Работа в выходные включена.", False),
    ("Не включено / без ставки: раскадровка (раскадровщица агентства), диктор с правами, композитор с правами - строки со знаком «?»; глобальные правки после фиксации этапа и доп. раунды - отдельно; юр. согласование Сбера - на этапе аниматика.", False),
    ("Ставки: рейт-карта студии от 18.08.2026 (лист EXPENSES); комиссия студии и налоги - инпуты D22/D23, по умолчанию 5% / 15% как в смете VTB Privilege 18.08 (в шаблоне 14.08 были 10% / 15%).", False),
]
for i, (t, b) in enumerate(notes):
    c = m.cell(row=29 + i, column=2); c.value = t
    c.font = Font(name="Roboto", bold=b, sz=10 if b else 9)
    c.alignment = Alignment(wrap_text=False)

wb.save(OUT)
print("saved", OUT)
