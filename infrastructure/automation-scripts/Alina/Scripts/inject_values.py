"""Compute cached values for the simple formulas in the estimate and inject <v> into sheet XML
(no LibreOffice on this Mac). Formulas used: sheet refs, SUM(range), + - * / ( ), $ anchors."""
import re, zipfile, shutil, sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
SRC=sys.argv[1]; OUT=sys.argv[2]
wb=load_workbook(SRC)
memo={}
def val(sheet,coord):
    key=(sheet,coord)
    if key in memo: return memo[key]
    v=wb[sheet][coord].value
    if isinstance(v,str) and v.startswith("="): r=evalf(sheet,v[1:])
    elif isinstance(v,(int,float)): r=float(v)
    else: r=0.0
    memo[key]=r; return r
def rng_sum(sheet,rng):
    c1,r1,c2,r2=range_boundaries(rng); s=0.0
    for r in range(r1,r2+1):
        for c in range(c1,c2+1): s+=val(sheet,f"{get_column_letter(c)}{r}")
    return s
def evalf(sheet,f):
    f=f.replace("$","")
    def sumrep(m):
        sh,rng=m.group(1),m.group(2); sh=(sh or sheet).strip("'!")
        return repr(rng_sum(sh,rng))
    f=re.sub(r"SUM\(((?:'[^']+'!|[A-Za-z][A-Za-z ]*!)?)([A-Z]+\d+:[A-Z]+\d+)\)",sumrep,f)
    def refrep(m):
        sh,coord=m.group(1),m.group(2); sh=(sh or sheet).strip("'!")
        return repr(val(sh,coord))
    f=re.sub(r"((?:'[^']+'!|[A-Za-z][A-Za-z ]*!)?)([A-Z]{1,2}\d+)",refrep,f)
    return float(eval(f,{"__builtins__":{}},{}))
# compute all
values={}
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value.startswith("="):
                values[(ws.title,c.coordinate)]=val(ws.title,c.coordinate)
# map sheet name -> xml path
z=zipfile.ZipFile(SRC)
wbxml=z.read("xl/workbook.xml").decode(); rels=z.read("xl/_rels/workbook.xml.rels").decode()
rid={m.group(1):m.group(2) for m in re.finditer(r'<Relationship[^>]*Id="([^"]+)"[^>]*Target="([^"]+)"',rels)}
rid.update({m.group(2):m.group(1) for m in re.finditer(r'<Relationship[^>]*Target="([^"]+)"[^>]*Id="([^"]+)"',rels)})
sheets={}
for m in re.finditer(r'<sheet [^>]*name="([^"]+)"[^>]*r:id="([^"]+)"',wbxml):
    name=m.group(1).replace("&amp;","&"); target=rid[m.group(2)]
    sheets[name]="xl/"+target if not target.startswith("/") else target[1:]
out=zipfile.ZipFile(OUT,"w",zipfile.ZIP_DEFLATED)
for item in z.infolist():
    data=z.read(item.filename)
    for name,path in sheets.items():
        if item.filename==path:
            x=data.decode()
            def cellrep(m):
                coord=m.group(1); body=m.group(0)
                if (name,coord) in values and "<f>" in body:
                    v=f"<v>{values[(name,coord)]!r}</v>"
                    if "<v/>" in body: body=body.replace("<v/>",v)
                    elif "<v></v>" in body: body=body.replace("<v></v>",v)
                    elif "<v>" not in body: body=body.replace("</f>","</f>"+v)
                return body
            x=re.sub(r'<c r="([A-Z]+\d+)"[^>]*>.*?</c>',cellrep,x,flags=re.S)
            data=x.encode()
    out.writestr(item,data)
out.close()
print("MAIN E21 subtotal:",values[("MAIN","E21")]," E24 total:",values[("MAIN","E24")])
for k in [("PRE PRODUCTION","K14"),("PRE PRODUCTION","K20"),("PRODUCTION","K17"),("POST PRODUCTION","K13"),("POST PRODUCTION","K23"),("POST PRODUCTION","K30"),("EXPENSES","F14")]:
    if k in values: print(k, round(values[k],2))
