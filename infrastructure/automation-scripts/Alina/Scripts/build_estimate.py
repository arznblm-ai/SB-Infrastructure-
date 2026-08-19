from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from copy import copy
from openpyxl.cell.cell import MergedCell

SRC="template.xlsx"; OUT="EST_TimeFrame_3D_animation_18-08-26.xlsx"
FX=85.0136  # CBR official USD/RUB, 18.08.2026
wb=load_workbook(SRC)

def en(s):
    """strip ' / Russian' tail"""
    if isinstance(s,str) and " / " in s: return s.split(" / ")[0].strip()
    return s

# ---------------- EXPENSES ----------------
ex=wb["EXPENSES"]
card={8:15000,9:25000,10:30000,11:20000,12:25000,           # main crew
      14:25000,15:15000,16:20000,17:20000,18:25000,19:15000,20:25000,21:20000,  # 3D
      23:15000,24:20000,25:20000,26:20000,27:15000,28:18000,29:15000,30:30000}  # 2D & post
ex["B7"]="Main crew's expenses"; ex["B13"]="3D Department expenses"; ex["B22"]="2D & Post-prod Department expenses"
# header row 6: unmerge and label
for rng in list(ex.merged_cells.ranges):
    if str(rng)=="B6:E6": ex.unmerge_cells("B6:E6")
ex["C6"]="Role"; ex["E6"]="RUB / day"; ex["F6"]="USD / day"
for c in ("C6","E6","F6"):
    ex[c].font=Font(name="Roboto",bold=True); ex[c].alignment=Alignment(horizontal="center")
for r in range(6,31):
    src=ex.cell(row=r,column=5); dst=ex.cell(row=r,column=6)
    dst.border=copy(src.border); dst.font=copy(src.font); dst.alignment=copy(src.alignment)
    dst.number_format="#,##0.00"
for r,v in card.items():
    ex.cell(row=r,column=3).value=en(ex.cell(row=r,column=3).value)
    ex.cell(row=r,column=5).value=v
    ex.cell(row=r,column=5).number_format="#,##0"
    ex.cell(row=r,column=6).value=f"=E{r}/MAIN!$F$27"
ex.column_dimensions["F"].width=14
ex["B32"]="Rates: studio rate card (RUB per day); USD at the exchange rate on MAIN. Overtime not planned for this project."
ex["B32"].font=Font(name="Roboto",italic=True,sz=9)

# ---------------- helper for detail sheets ----------------
def clean_sheet(ws,title):
    ws["C2"]=title
    for rng in list(ws.merged_cells.ranges):
        if rng.min_col==3 and rng.max_col==4 and rng.min_row==rng.max_row:
            ws.unmerge_cells(str(rng))   # C:D merges swallow the Note column
    for col in "CDEFGHIJK": ws[f"{col}5"]=None      # Russian header row
    ws["G4"]="Day rate (USD)"; ws["K4"]="Total (USD)"
    for row in ws.iter_rows(min_row=6,max_row=ws.max_row):
        for c in row:
            if isinstance(c,MergedCell): continue
            if c.column==3 and isinstance(c.value,str): c.value=en(c.value)
            if c.column==9: c.value=None                     # OT rate (RUB) - clear
            if c.column==7 and isinstance(c.value,str) and c.value.startswith("=EXPENSES!E"):
                c.value=c.value.replace("=EXPENSES!E","=EXPENSES!F")
            if c.column==7 and isinstance(c.value,(int,float)): c.value=None   # hardcoded RUB rates
            if c.column in (5,8) and c.row>=6 and not (isinstance(c.value,str) and c.value.startswith("=")):
                c.value=None                                  # reset q-ty / days
            if c.column in (7,11): c.number_format="#,##0.00"

def line(ws,r,note=None,qty=None,days=None,label=None):
    if label is not None: ws.cell(row=r,column=3).value=label
    if note is not None:
        ws.cell(row=r,column=4).value=note
        ws.cell(row=r,column=4).alignment=Alignment(wrap_text=True,vertical="top")
    ws.cell(row=r,column=5).value=qty
    ws.cell(row=r,column=8).value=days

# ---------------- PRE PRODUCTION ----------------
pp=wb["PRE PRODUCTION"]; clean_sheet(pp,"PRE PRODUCTION")
pp["B7"]="Main crew's expenses"; pp["B16"]="Preproduction expenses"; pp["B22"]="Tests"
line(pp,8,"Coordination, approvals (look-dev gate, v1), client communication, delivery",1,5)
line(pp,9,"Look-dev direction: walnut tone, groove ring read, lighting; framing in square / 16:9 / 9:16; review of stills and v1",1,3)
line(pp,17,"n/a - creative brief, shot list and references supplied by client",label="Creative development")
line(pp,18,"n/a - shots defined in the brief",label="Storyboard")
line(pp,23,"n/a - CAD, geometry plates, brand assets and screen content supplied by client")
line(pp,24,"Frame-time / render test - included in look-dev",label="Tests")
pp["E25"]=None

# ---------------- PRODUCTION ----------------
pr=wb["PRODUCTION"]; clean_sheet(pr,"PRODUCTION")
pr["B7"]="3D scene assembly, look-dev & animation"; pr["B19"]="Shooting"
line(pr,8,"CAD prep; master materials (real-scale African walnut, black matte stepped groove ring, matte non-emissive e-ink); lighting per shot; 8 look-dev stills at final quality; render supervision (beauty, screen pass, alpha); revision round",1,10)
line(pr,9,"CAD clean-up, tessellation for macro shots, 2 updated geometry statics (Small / Large)",1,1)
line(pr,11,"included in 3D designer")
line(pr,14,"Camera, light and focus animation for 8 renders (7 shots + Large variant of Shot 2): whip rotation with drifting holds, seamless loops, macro slider with focus pulls, pedestal arc, fly-around, size comparison, push-in; low-res previews",1,5)
for r in (20,21,22,23):
    pr.cell(row=r,column=11).value=f"=E{r}*(G{r}*H{r}+I{r}*J{r})"
line(pr,20,"n/a - full CGI, no shooting")

# ---------------- POST PRODUCTION ----------------
po=wb["POST PRODUCTION"]; clean_sheet(po,"POST PRODUCTION")
po["B7"]="Edit & mastering"; po["B15"]="CGI & VFX (compositing)"; po["B25"]="Sound"
line(po,11,"included - both crops from the 2160 x 2160 square master",label="Resizes")
line(po,16,label="Retoucher")
line(po,18,"Screen pass compositing (client PNG content on the matte e-ink panel, lit by scene light); mid-shot screen changes on Shots 5 & 7; 16:9 and 9:16 crops; mastering & exports per shot: ProRes 4444 (+alpha), PNG sequence, H.264 review, separate screen pass",1,7)
line(po,19,"planar screen tracking - included in compositing")
line(po,26,"n/a - no audio per brief")

# ---------------- MAIN ----------------
m=wb["MAIN"]
m["C6"]="TimeFrame"
m["C7"]="TimeFrame - 3D product animation (2 updated statics + 7 shots)"
m["B8"]="Date:"; m["C8"]="August 18, 2026"
m["E9"]="USD"; m["F9"]=None
m["B10"]="PRE PRODUCTION"; m["B11"]="Main crew's expenses"; m["B12"]="Preproduction expenses"; m["B13"]="Tests"
m["B14"]="PRODUCTION"; m["B15"]="3D scene assembly, look-dev & animation"; m["B16"]="Shooting"
m["B17"]="POST PRODUCTION"; m["B18"]="Edit & mastering"; m["B19"]="CGI & VFX (compositing)"; m["B20"]="Sound"
m["B21"]="SUBTOTAL:"; m["B22"]="Production Fee"; m["B23"]="Taxes"; m["B24"]="TOTAL:"
refs={11:"='PRE PRODUCTION'!K14",12:"='PRE PRODUCTION'!K20",13:"='PRE PRODUCTION'!K26",
      15:"=PRODUCTION!K17",16:"=PRODUCTION!K24",
      18:"='POST PRODUCTION'!K13",19:"='POST PRODUCTION'!K23",20:"='POST PRODUCTION'!K30"}
for r,f in refs.items():
    m.cell(row=r,column=5).value=f
for r in range(11,25):
    if not isinstance(m.cell(row=r,column=6),MergedCell): m.cell(row=r,column=6).value=None
    if not isinstance(m.cell(row=r,column=5),MergedCell): m.cell(row=r,column=5).number_format="#,##0.00"
m["E21"]="=SUM(E11:E20)"
m["D22"]=0.0; m["E22"]="=E21*D22"
m["D23"]=0.0; m["E23"]="=(E21+E22)*D23"
m["E24"]="=E21+E22+E23"
m["E27"]=None
m["B27"]="USD / RUB exchange rate (Bank of Russia, 18 Aug 2026):"
m["F27"]=FX; m["F27"].number_format="#,##0.00"
notes=[("Notes",True),
 ("Scope: 2 updated geometry statics (Small / Large) and 7 animated shots per the brief (Shot 2 delivered in Small and Large), 30 fps, no audio.",False),
 ("Deliverables per shot: square master 2160 x 2160 min cropped to 16:9 and 9:16; ProRes 4444 (+alpha where applicable), full PNG sequence, H.264 review copy; screen pass delivered separately; project source files.",False),
 ("Process: one look-dev still per shot at final quality approved before any animation renders (hard gate); one revision round included after look-dev approval.",False),
 ("Not included: global changes after a stage is approved and additional revision rounds (quoted separately); music / voice-over; KeyShot port of the scenes (available as an option).",False),
 ("Timeline: approx. 5 weeks from asset handoff, subject to look-dev and v1 approvals within 2-3 business days each.",False),
 ("Rates: studio rate card (see EXPENSES) converted at the exchange rate above; production fee and taxes are set to 0% for this estimate.",False)]
for i,(t,b) in enumerate(notes):
    c=m.cell(row=29+i,column=2); c.value=t; c.font=Font(name="Roboto",bold=b,sz=9 if not b else 10)
wb.save(OUT); print("saved",OUT)
